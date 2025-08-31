# final_teams_title.py
# -*- coding: utf-8 -*-
import requests
from bs4 import BeautifulSoup, FeatureNotFound
import re
import time
import pandas as pd
import os
import gzip
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
SITEMAP = "https://teams-titles.hr.ufl.edu/sitemap.xml"
MAX_WIDTH = 50
REQUEST_TIMEOUT = 30
SLEEP_BETWEEN = 0.15  # cortesía entre requests

# ---------- HELPERS ----------
def _get_request_content(url):
    """Descarga contenido y maneja .gz si aplica."""
    r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    content = r.content
    # detect gzip by magic bytes or extension
    if url.lower().endswith(".gz") or (len(content) >= 2 and content[:2] == b"\x1f\x8b"):
        try:
            with gzip.GzipFile(fileobj=io.BytesIO(content)) as gf:
                return gf.read()
        except Exception:
            return content
    return content

def extract_section_from_container(container, title_patterns, stop_patterns=None):
    if not container:
        return ""
    header = container.find(
        lambda t: t.name in ["h2", "h3"]
        and t.get_text()
        and any(re.search(p, t.get_text(), re.I) for p in title_patterns)
    )
    if not header:
        return ""
    parts = []
    for sib in header.next_siblings:
        name = getattr(sib, "name", None)
        if name in ["h2", "h3"]:
            break
        text = sib.get_text(" ", strip=True) if getattr(sib, "get_text", None) else ""
        if stop_patterns and text and any(re.search(sp, text, re.I) for sp in stop_patterns):
            break
        if name == "ul":
            for li in sib.find_all("li", recursive=False):
                li_text = li.get_text(" ", strip=True)
                if li_text:
                    parts.append("♦ " + li_text)
        else:
            if text:
                parts.append(text)
    return "\n".join(parts).strip()

def find_in_main(container_text, patterns, stop_words=None):
    for p in patterns:
        m = re.search(p, container_text, re.I)
        if m and m.group(1):
            val = m.group(1).strip()
            if stop_words:
                for sw in stop_words:
                    val = re.split(sw, val, flags=re.I)[0]
            return val.strip()
    return ""

def extract_job_from_url(url):
    """Extrae los campos de un job a partir de su URL (robusto para páginas individuales)."""
    r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    main = soup.find('div', class_=lambda c: c and ("entry-content" in c or "content" in c or "post-content" in c))
    if not main:
        main = soup.find('main') or soup

    main_text = main.get_text(" ", strip=True).replace("\xa0", " ")
    main_text = re.sub(r"\s+", " ", main_text)

    h1 = soup.find("h1") or (main.find("h1") if main else None)
    title = h1.get_text(strip=True) if h1 else ""

    job_code    = find_in_main(main_text, [r"Job\s*Code[:\s\-]*([0-9]{3,})"])
    flsa_status = find_in_main(main_text, [r"FLSA\s*Status[:\s\-]*([A-Za-z\s\-]+)"], stop_words=["Pay"])
    pay_grade   = find_in_main(main_text, [r"Pay\s*Grade[:\s\-]*([0-9A-Za-z\-]+)"])
    summary     = extract_section_from_container(main, [r"^\s*Summary\s*$", r"\bSummary\b"], stop_patterns=["Examples of Work","Education","Licensure","Supervision","Job Families","Competencies"])
    examples    = extract_section_from_container(main, [r"Examples\s+of\s+Work"], stop_patterns=["Education","Licensure","Supervision","Job Families","Competencies"])
    education   = extract_section_from_container(main, [r"Education\s+and\s+Experience"], stop_patterns=["Licensure","Supervision","Job Families","Competencies"])
    licensure   = extract_section_from_container(main, [r"Licensure\s+and\s+Certification"], stop_patterns=["Supervision","Job Families","Competencies"])
    supervision = extract_section_from_container(main, [r"^Supervision$","\bSupervision\b"], stop_patterns=["Job Families","Competencies"])
    competencies = extract_section_from_container(main, [r"Competencies"], stop_patterns=["Job Families"])
    if not competencies:
        competencies = "not found"

    return {
        "Job Title": title,
        "URL": url,
        "Job Code": job_code or "",
        "FLSA Status": flsa_status or "",
        "Pay Grade": pay_grade or "",
        "Summary": summary or "",
        "Examples of Work": examples or "",
        "Education and Experience": education or "",
        "Licensure and Certification": licensure or "",
        "Supervision": supervision or "",
        "Competencies": competencies
    }

def _parse_sitemap_urls(sitemap_url):
    """Extrae URLs '/teams-title/' desde sitemap principal o subsitemaps (.xml y .xml.gz)."""
    try:
        raw = _get_request_content(sitemap_url)
    except Exception as e:
        print("Error fetching sitemap:", e)
        return []
    try:
        smap = BeautifulSoup(raw, "xml")
    except Exception:
        smap = BeautifulSoup(raw, "html.parser")

    locs = [loc.get_text(strip=True) for loc in smap.find_all("loc")]
    urls = []
    for u in locs:
        if u.lower().endswith((".xml", ".xml.gz")) and "sitemap" in u.lower():
            try:
                sub_raw = _get_request_content(u)
                subsoup = BeautifulSoup(sub_raw, "xml")
                for loc in subsoup.find_all("loc"):
                    tu = loc.get_text(strip=True)
                    if "/teams-title/" in tu:
                        urls.append(tu)
            except Exception:
                continue
        elif "/teams-title/" in u:
            urls.append(u)
    return sorted(set(urls))

def _format_excel(file_path):
    """Formatea el Excel usando openpyxl y guarda como *_formatted.xlsx"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    bold_font = Font(bold=True, color="000000")
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(wrap_text=True, vertical="top")

    # Encabezados
    header_row = 1
    for cell in ws[header_row]:
        cell.font = bold_font
        cell.fill = gray_fill
        cell.border = thin_border
        cell.alignment = header_alignment
    try:
        ws.row_dimensions[header_row].height = 24
    except Exception:
        pass

    # Celdas cuerpo
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                cell.border = thin_border
                cell.alignment = cell_alignment

    # Ancho columnas
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value:
                lines = str(cell.value).splitlines()
                longest = max(len(line) for line in lines) if lines else len(str(cell.value))
                if longest > max_len:
                    max_len = longest
        adjusted = min(max_len + 2, MAX_WIDTH)
        if adjusted < 5:
            adjusted = 5
        ws.column_dimensions[col_letter].width = adjusted

    base, ext = os.path.splitext(file_path)
    salida = f"{base}_formatted{ext}"
    wb.save(salida)
    return salida

def _decide_urls(input_url: str | None):
    """Decide lista de URLs a procesar: sitemap completo si root/empty/sitemap, o lista con la URL individual."""
    if not input_url or not input_url.strip():
        return _parse_sitemap_urls(SITEMAP)
    url = input_url.strip()
    # treat the root domain as "all"
    root_variants = ("https://teams-titles.hr.ufl.edu", "https://teams-titles.hr.ufl.edu/")
    if url in root_variants:
        return _parse_sitemap_urls(SITEMAP)
    if url.endswith(".xml") or "sitemap" in url.lower():
        return _parse_sitemap_urls(url)
    # if it looks like a teams-title page, handle as single
    return [url]

# ---------- PUBLIC API ----------
def run_scraping(input_url: str | None = None, progress_cb=None):
    """
    Ejecuta scraping. progress_cb(curr:int, total:int, msg:str) es opcional.
    Retorna la ruta al archivo EXCEL ya formateado (ej: uf_jobs_full_formatted.xlsx).
    """
    urls = _decide_urls(input_url)
    if not urls:
        return None

    rows = []
    total = len(urls)
    for i, url in enumerate(urls, start=1):
        if progress_cb:
            progress_cb(i, total, f"Processing {i} of {total}")
        try:
            rows.append(extract_job_from_url(url))
        except Exception as e:
            # agrega una fila con error visible
            rows.append({
                "Job Title": "",
                "URL": url,
                "Job Code": "",
                "FLSA Status": "",
                "Pay Grade": "",
                "Summary": f"ERROR: {e}",
                "Examples of Work": "",
                "Education and Experience": "",
                "Licensure and Certification": "",
                "Supervision": "",
                "Competencies": ""
            })
        time.sleep(SLEEP_BETWEEN)
        if progress_cb:
            progress_cb(i, total, f"Processed {i} of {total}")

    df = pd.DataFrame(rows)
    base = "uf_job_single" if total == 1 else "uf_jobs_full"
    xlsx_path = f"{base}.xlsx"
    csv_path = f"{base}.csv"

    df.to_excel(xlsx_path, index=False)
    df.to_csv(csv_path, index=False)

    formatted = _format_excel(xlsx_path)

    # rename formatted to exactly the names you requested
    if total == 1:
        target = "uf_job_single_formatted.xlsx"
    else:
        target = "uf_jobs_full_formatted.xlsx"
    # move/rename if necessary
    if os.path.exists(formatted):
        os.replace(formatted, target)
        return target
    else:
        # fallback: if formatting failed, return raw
        return xlsx_path

if __name__ == "__main__":
    # quick manual test (careful: this will try to fetch all)
    out = run_scraping(None)
    print("Done:", out)
